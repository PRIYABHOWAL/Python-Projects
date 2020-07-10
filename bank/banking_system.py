import random
import pandas as pd
import string
import openpyxl as op
def excel():
    workbook = op.load_workbook('details.xlsx')
    sheet =workbook['Sheet1']
    row = sheet.max_row
    return(workbook,sheet,row)
def create_account():
    acc_holder_name=input("ENTER YOUR NAME::")
    letters=string.ascii_letters
    numbers=string.digits
    password=''.join(random.choice(letters+numbers) for i in range(10))
    acc_no=random.randint(100000,1000000)
    amount=0
    print("---------HELLO"+" "+str(acc_holder_name.upper())+"---------")
    print("YOUR ACCOUNT NO IS"+" " +str(acc_no))
    print("YOUR PASSWORD IS"+" "+str(password))
    print("ACCOUNT CREATED SUCCESSFULLY")
    details={"account":[acc_no],"name":[acc_holder_name],"password":[password],"amount":[amount]}
    df1=pd.DataFrame.from_dict(details)
    book = op.load_workbook('details.xlsx')
    sheet=book["Sheet1"]
    writer = pd.ExcelWriter('details.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    for sheetname in writer.sheets:
        df1.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)
    writer.save()
_acc = 1
_pas = 3
_name = 2
_amount = 4
def login():
    logins = (False,0)
    acc_no=input("ENTER YOUR ACCOUNT NO::")
    passw=input("ENTER THE PASSWORD::")
    workbook,sheet,row = excel()
    for i in range(1,row+1):
        acc = sheet.cell(i,1).value
        password = sheet.cell(i,_pas).value
        if acc_no == str(acc):
            if passw == str(password):
                print("Login Successful!")
                logins=(True,i)
                def amount(logins):
                    workbook,sheet,row = excel()
                    amounts = sheet.cell(logins[1],_amount).value
                    print("YOU have"+" "+ "Rs"+" "+str(amounts)+" " +"in your account")
                amount(logins)
        
    if logins[0]==True:
        while True:
            choice =int(input('Select option \n 1. Deposit Money\n 2. Withdraw Money \n 3. Transfer Money\n 4. Delete Account\n 5. exit\n'))
            if choice==1:
                def deposit(logins):
                    amount_add=int(input("ENTER AMOUNT YOU WANT TO ADD::"))
                    workbook,sheet,row = excel()
                    sheet.cell(logins[1],_amount).value+=amount_add
                    workbook.save("details.xlsx")
                    print("Rs"+" "+str(amount_add)+" " +"added successfully.")
                    amount(logins)
                deposit(logins)
            elif(choice==2):
                def withdraw(logins):
                    workbook,sheet,row = excel()
                    amount_remove=int(input("ENTER AMOUNT YOU WANT TO WITHDRAW"))
                    if sheet.cell(logins[1],_amount).value <amount_remove:
                        print("UNSUFFICIENT BALANCE IN ACCOUNT")
                    else:
                        sheet.cell(logins[1],_amount).value-=amount_remove
                        workbook.save("details.xlsx")
                        print("Rs"+" "+str(amount_remove)+" " +" withdrawn successfully.")
                        amount(logins)
                withdraw(logins)
            elif(choice==3):
                def transfer(logins):
                    workbook,sheet,row = excel()
                    another_acc_no=int(input('ENTER THE ACCOUNT NO U WANT TO TRANSFER MONEY'))
                    for i in range(1,row+1):
                        if another_acc_no==sheet.cell(i,_acc).value:
                            amount_transfer=int(input("ENTER THE AMOUNT U WANT TO TRANSFER"))
                            if sheet.cell(logins[1],_amount).value <amount_transfer:
                                print("UNSUFFICIENT BALANCE IN ACCOUNT")
                            else:
                                sheet.cell(logins[1],_amount).value-=amount_transfer
                                sheet.cell(i,_amount).value+=amount_transfer
                                workbook.save("details.xlsx")
                                print("Rs"+" "+str(amount_transfer)+" " +" transferred successfully.")
                                amount(logins)
                transfer(logins)
            elif(choice==4):
                def delete(logins):
                    workbook,sheet,row = excel()
                    for i in range(1,row+1):
                        acc = sheet.cell(i,1).value
                        password = sheet.cell(i,_pas).value
                        if acc_no == str(acc):
                            if passw == str(password):
                                sheet.cell(i,1).value=None
                                sheet.cell(i,2).value=None
                                sheet.cell(i,3).value=None
                                sheet.cell(i,4).value=None
                                print("DELETED SUCCESSFUllY")
                                workbook.save("details.xlsx")
                delete(logins)
            elif(choice==5):
                print("THANK YOU")
                break
            else:
                print("ENTER A VALID CHOICE")
                break
print("=============================")
print('\tWELCOME TO BANK\t')
print("=============================")
while True:
    n=int(input('Select option \n 1. Create Account \n 2. Login Account \n 3. exit\n'))
    if(n==1):
        create_account()  
    elif(n==2):
        login()
    elif(n==3):
        print("THANK YOU")
    else:
        print("ENTER A VALID CHOICE")
